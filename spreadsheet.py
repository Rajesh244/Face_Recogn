from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import time,datetime
import os,sys
import sqlite3

excel="F:/Autoattendance-Cognitive-master-Copy/excel"
csv="F:/Autoattendance-Cognitive-master-Copy/csv"
if not os.path.exists(excel):
    os.makedirs(excel)

if not os.path.exists(csv):
    os.makedirs(csv)
    
classid="IT2015"#sys.argv[1]
#database connection
conn = sqlite3.connect('Face-DataBase')
c = conn.cursor()
conn1=sqlite3.connect('test')
t=conn1.cursor()
#get current date
currentDate = time.strftime("%d_%m_%y")
d = datetime.datetime.today()
current_mon=d.strftime("%B-%Y")
current_day=d.strftime("%A")
t.execute("select * from "+classid+"TT where day='%s'" %current_day)
row=t.fetchall()
row=list(row[0])
#create a workbook and add a worksheet
if(os.path.exists('./excel/'+classid+'.xlsx')):
    wb = load_workbook(filename = "F:/Autoattendance-Cognitive-master-Copy/excel/"+classid+".xlsx")
    if current_mon in wb.sheetnames:
        sheet = wb[current_mon]
    else:
        sheetx=wb.active
        z=3
        day_count=0
        while(1):
            if sheetx.cell(row=2,column=z).value is None:
                sheetx.cell(row=2,column=z).value="Attendace Percentage"
                break
            else:
                z=z+7
                day_count+=1
                #print(get_column_letter(z))
                #print(day_count)
        (r,cl)=(3,3)
        
        while(1):
            (absent,cl)=(0,3)
            if sheetx.cell(row=r,column=1).value is not None:
                while(1):
                    if sheetx.cell(row=r,column=cl).value=="AB" and sheetx.cell(row=r,column=cl).value is not None:
                        absent+=1
                        cl+=1
                    else:
                        percent=100-((absent/day_count)*100)
                        sheetx.cell(row=r,column=z).value=percent
                        break
                r+=1
            else:
                break
        (r1,c1,r2,c2)=(1,3,1,9)
        c.execute("SELECT * FROM "+classid+" ORDER BY Roll ASC")
    
    #creating worksheet and giving names to column
        ws1 = wb.create_sheet(current_mon)
        #ws1.title = current_mon
        ws1.append(('RollNumber','Name'))
        ws1.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)
        ws1.cell(row=r1,column=c1).value=currentDate
        ws1.append({'C':row[1],'D':row[2],'E':row[3],'F':row[4],'G':row[5],'H':row[7],'I':row[8]})
        #entering students information from database
        while True:
            a = c.fetchone()
            if a == None:
                break
            else:
                ws1.append((a[2], a[1]))

    #saving the file
        wb.save(filename = "F:/Autoattendance-Cognitive-master-Copy/excel/"+classid+".xlsx")

    # sheet[ord() + '1']
    for col_index in range(1, 850):
    	col = get_column_letter(col_index)
    	if sheet['%s%s' % (col,1)].value is None:
            col2 = get_column_letter(col_index-1)
            if sheet['%s%s' % (col2,1)].value != currentDate:
                c1=col_index+6
                c2=col_index+12
                r1=1
                r2=1
                sheet.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)
                sheet.cell(row=r1,column=c1).value=currentDate
                
                sheet.cell(row=2,column=c1).value=row[1]
                sheet.cell(row=2,column=c1+1).value=row[2]
                sheet.cell(row=2,column=c1+2).value=row[3]
                sheet.cell(row=2,column=c1+3).value=row[4]
                sheet.cell(row=2,column=c1+4).value=row[5]
                sheet.cell(row=2,column=c1+5).value=row[7]
                sheet.cell(row=2,column=c1+6).value=row[8]
                wb.save(filename = "F:/Autoattendance-Cognitive-master-Copy/excel/"+classid+".xlsx")
            break

    #saving the file
    wb.save(filename ="F:/Autoattendance-Cognitive-master-Copy/excel/"+ classid+".xlsx")
    	
else:
    wb = Workbook()
    dest_filename = "F:/Autoattendance-Cognitive-master-Copy/excel/"+classid+'.xlsx'
    c.execute("SELECT * FROM "+classid+" ORDER BY Roll ASC")
    
    #creating worksheet and giving names to column
    ws1 = wb.active
    ws1.title = current_mon
    ws1.append(('RollNumber','Name'))
    ws1.merge_cells(start_row=1,start_column=3,end_row=1,end_column=9)
    ws1.cell(row=1,column=3).value=currentDate
    ws1.append({get_column_letter(3):row[1],get_column_letter(4):row[2],get_column_letter(5):row[3],get_column_letter(6):row[4],get_column_letter(7):row[5],get_column_letter(8):row[7],get_column_letter(9):row[8]})
    
    #entering students information from database
    while True:
        a = c.fetchone()
        if a == None:
            break
        else:
            ws1.append((a[2], a[1]))

    #saving the file
    wb.save(filename = dest_filename)