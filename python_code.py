import os
from flask import Flask, render_template, request, redirect, url_for,session
from datetime import timedelta
import sqlite3,time
from flask_wtf import Form
from wtforms import SubmitField,SelectField
import xlrd
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

currentDate = time.strftime("%d_%m_%y")
current_mon=time.strftime("%B-%Y")

#Web application
app = Flask(__name__)
app.secret_key='qwerty'
@app.route('/')
def index():
   #print(session)
    return render_template('faculty_login.html')

@app.route('/login', methods = ['POST', 'GET'])   #endpoint that is called from the html template when button is pressed
def displayNames():
    if request.method== 'POST':
        fid = request.form['id']
        fpassword = request.form['password']
        conn = sqlite3.connect('test')
        try:
            c = conn.cursor()
            sql = "SELECT * FROM FACULTY WHERE ID='%s'" % (fid)
            c.execute(sql)
            results = c.fetchall()
            isRecordExist = 0
            for row in results:                                                          # checking wheather the id exist or not
                isRecordExist = 1
            if isRecordExist == 1:
                for row in results:
                    id = row[1]
                    password = row[2]
            else:
                print("Invalid ID")
                return redirect(url_for('index'))
        except:
            print("Error: unable to fecth data")
            return redirect(url_for('index'))
        if fid != id or fpassword != password :
            print("Invalid credentials.Try again")
            return redirect(url_for('index'))
        else:
            session['userid']=id
            #print(session)
            return redirect(url_for('view'))
        conn.commit()
        conn.close()
    else:
        return redirect(url_for('index'))
@app.route('/view',methods = ['POST','GET'])

def view():
    if 'userid' in session:
        userid=session['userid']
    else:
        return redirect(url_for('displayNames'))
#form
    class_list=set()
    conn=sqlite3.connect('test')
    c=conn.cursor()
    sql="select courseid,classname from course where id='%s'" %(userid)
    c.execute(sql)
    result=c.fetchall()
    sql="select incharge from faculty where id='%s'" %(userid)
    c.execute(sql)
    incharge=c.fetchall()
    for i in result:
            class_list.add((i[0]+" "+i[1],i[0]))
    for i in incharge:
        if i=="IT" or i=="CSE" or i=="ECE" or i=="EEE" or i=="MECH" :
            sql="select * from classname where id like '%s%'" %(userid)
            c.execute(sql)
            in_charge=c.fetchall()
            for j in in_charge:
                class_list.add((j[0],j[0]))
        else:
            class_list.add((i[0],i[0]))
#Form Class
    class View_form(Form):
        Class=SelectField('Class',choices=class_list)
        submit=SubmitField("View")
    form=View_form()
    if(request.method=='POST'):
        form_data=request.form['Class']
        if len(form_data) >7:
            course_id=form_data[:6]
            class_id=form_data[7:]
        else:
            class_id=form_data
        wb = xlrd.open_workbook("F:/Autoattendance-Cognitive-master-Copy/excel/"+class_id+'.xlsx')
        x='''<form action="{{url_for('month')}}" method="POST">
        <input type="text" name='Class' value="'''+form_data+'''"readonly><br><select name="month">
        '''
        for month in wb.sheet_names():
            x+='''<option value="'''+month+'''">'''+month+'''</option>'''
        x+='''</select><input type="submit" value="View"></form>'''
        f=open("F:/Autoattendance-Cognitive-master-Copy/templates/view_month.html",'w')
        f.write(x)
        f.close()
        return render_template("view_month.html")
    else:
        return render_template("view_form.html",form=form)
    
@app.route('/month',methods=["POST","GET"])
def month():
    if request.method=='POST':
        form_data=request.form['Class']
        form_mon=request.form['month']
        if len(form_data) >7:
            course_id=form_data[:6]
            class_id=form_data[7:]
            wb = xlrd.open_workbook("F:/Autoattendance-Cognitive-master-Copy/excel/"+class_id+'.xlsx')
            #if current_mon in wb.sheetnames:
            sh = wb.sheet_by_name(form_mon)

          #  os.remove("E:/Autoattendance-Cognitive-master-Copy/csv/"+course_id+"_"+class_id+'.csv')
            your_csv_file = open("F:/Autoattendance-Cognitive-master-Copy/csv/"+course_id+"_"+class_id+'.csv', 'w', encoding='utf8')
            wr = csv.writer(your_csv_file, quoting=csv.QUOTE_ALL)

            for rownum in range(sh.nrows):
                wr.writerow(sh.row_values(rownum))

            your_csv_file.close()
            m=course_id
            def csv_to_html_table(fname,headers=None,delimiter=","):
                with open(fname) as f:
                    content = f.readlines()
                    #reading file content into list
                    rows = [x.strip() for x in content]
                    table = "<table border='1'>"
                    #creating HTML header row if header is provided 
                    if headers is not None:
                        table+= "".join(["<th>"+cell[1:-1]+"</th>" for cell in headers.split(delimiter)])
                    else:
                        for cell in rows[0].split(delimiter):
                            if len(cell)>4:
                                table+= "".join(["<th>"+cell[1:-1]+"</th>"])
                        rows=rows[1:]
                    #Converting csv to html row by row
                    i=0
                    index=[0,1]
                    for cell in rows[1].split(delimiter):
                        if cell[1:-1]==m:
                            index.append(i)
                        i=i+1
                    rows=rows[3::2]
                    for row in rows:
                        row=row.split(delimiter)
                        table+="<tr>"
                        for i in index:
                            data=row[i]
                            table+="".join(["<td>"+data[1:-1]+"</td>"])
                        table+="</tr>" + "\n"
                    table+="</table><br>"
                    return table

            x=csv_to_html_table("F:/Autoattendance-Cognitive-master-Copy/csv/"+course_id+"_"+class_id+'.csv')
          #  os.remove("E:/Autoattendance-Cognitive-master-Copy/templates/"+course_id+"_"+class_id+'.html')
            f=open("F:/Autoattendance-Cognitive-master-Copy/templates/"+course_id+"_"+class_id+".html",'w')
            f.write(x)
            if form_mon==current_mon:
                y='''<fieldset><legend>On-Duty</legend><form action="{{url_for('edit')}}" method='POST'><input type="text" name="classid" value="'''+form_data+'''"readonly><br>
Date :'''+currentDate+'''<br> Roll num:<input type="text" name="roll"><br>   Period Num :<input type="text" name="period" ><br>Attendance : On Duty <br> <input type="submit" value="submit" name="submit">
</form><br></fieldset>
<br>
<button onclick="myFunction()">Print this page</button>

<script>
function myFunction() {
    window.print();
}
</script>
<form action="{{url_for('logout')}}">
<input type="submit" name="Logout" value="Logout">
</form><a href="/view">Go back</a>'''
                f.write(y)
            f.close()
            return render_template(course_id+"_"+class_id+".html")
           
        else:
            class_id=form_data
            wb = xlrd.open_workbook("F:/Autoattendance-Cognitive-master-Copy/excel/"+class_id+'.xlsx')

            #if current_mon in wb.sheetnames:
            sh = wb.sheet_by_name(form_mon)
            your_csv_file = open("F:/Autoattendance-Cognitive-master-Copy/csv/"+class_id+'.csv', 'w', encoding='utf8')
            wr = csv.writer(your_csv_file, quoting=csv.QUOTE_ALL)

            for rownum in range(sh.nrows):
                wr.writerow(sh.row_values(rownum))

            your_csv_file.close()

            def csv_to_html_table_main(fname,headers=None,delimiter=","):
                with open(fname) as f:
                    content = f.readlines()
                    #reading file content into list
                    rows = [x.strip() for x in content]
                    #print(rows)
                    table = "<table border='1'>"
                    #creating HTML header row if header is provided 
                    if headers is not None:
                        table+= "".join(["<th>"+cell[1:-1]+"</th>" for cell in headers.split(delimiter)])
                    else:
                        table+= "".join(["<th>"+cell[1:-1]+"</th>" for cell in rows[0].split(delimiter)])
                        rows=rows[1:]
                    #Converting csv to html row by row
                    for row in rows:
                        table+= "<tr>" + "".join(["<td>"+cell[1:-1]+"</td>" for cell in row.split(delimiter)]) + "</tr>" + "\n"
                    table+="</table><br>"
                    return table

            x=csv_to_html_table_main("F:/Autoattendance-Cognitive-master-Copy/csv/"+class_id+'.csv')
           # os.remove("E:/Autoattendance-Cognitive-master-Copy/templates/"+class_id+'.html')
            f=open("F:/Autoattendance-Cognitive-master-Copy/templates/"+class_id+".html",'w')
            f.write(x)
            if form_mon==current_mon:
                y='''<fieldset><legend>On-Duty</legend><form action="{{url_for('edit')}}" method='POST'><input type="text" name="classid" value="'''+class_id+'''"readonly><br>
Date :'''+currentDate+'''<br> Roll num:<input type="text" name="roll"><br>   Period Num :<input type="text" name="period" ><br>Attendance : On Duty <br> <input type="submit" value="submit" name="submit">
</form><br></fieldset>
<br>
<button onclick="myFunction()">Print this page</button>

<script>
function myFunction() {
    window.print();
}
</script>
<form action="{{url_for('logout')}}">
<input type="submit" name="Logout" value="Logout">
</form><a href="/view">Go back</a>'''
                f.write(y)
            #print(x)
            f.close()
            return render_template(class_id+'.html')
    else:
        return redirect(url_for('view')) 
    
@app.route('/s_add',methods=['POST','GET'])
def add_stud():
    
    if request.method== 'POST':
        s_name=request.form['sname']
        roll_num=request.form['rollnum']
        class_id=request.form['classid']
        os.system("python add_student.py "+s_name+" "+roll_num+" "+class_id)
    else:
        return render_template('add_student.html')
    
        

@app.route('/f_add',methods=['POST','GET'])
def add_fac():
    if request.method=='POST':
        f_name=request.form['username']
        password1=request.form['password1']
        password2=request.form['password2']
        id=request.form['id']
        if password1!=password2 :
            flash('Passwords do not match')
            return redirect(url_for('add_fac'))
        else:
            connect = sqlite3.connect('test')                                  # connecting to the database
            cmd = "SELECT * FROM FACULTY WHERE ID = '%s'" %(id)                             # selecting the row of an id into consideration
            cursor = connect.execute(cmd)
            isRecordExist = 0
            for row in cursor:                                                          # checking wheather the id exist or not
                isRecordExist = 1
                if isRecordExist == 1:                                                      # updating name and roll no
                    return redirect(url_for('add_fac'))
            params = (f_name,id,password1)                                               # insering a new student data
            connect.execute("INSERT INTO faculty(username,id,password) VALUES(?,?, ?)", params)
        connect.commit()                                                            # commiting into the database
        connect.close()                                                             # closing the connection
        return ' Faculty added successfully' + '''<br><a href='/f_add'>Click here to go back</a>'''
    else:
        return render_template('faculty_reg.html')

@app.route('/c_add',methods=['POST','GET'])
def add_class():
    if request.method=='POST':
            c_name=request.form['class_name']
            c_id=request.form['class_id']
            connect = sqlite3.connect('test')                                  # connecting to the database
            cmd = "SELECT * FROM CLASSNAME WHERE ID = '%s' "  %(c_name)                             # selecting the row of an id into consideration
            cursor = connect.execute(cmd)
            isRecordExist = 0
            for row in cursor:                                                          # checking wheather the id exist or not
                isRecordExist = 1
                if isRecordExist == 1:                                                      # updating name and roll no
                    return redirect(url_for('add_class'))
            params = (c_name,c_id)                                               # insering a new student data
            connect.execute("INSERT INTO classname(id,groupid) VALUES(?, ?)", params)
            connect.commit()                                                            # commiting into the database
            connect=sqlite3.connect('Face-DataBase')
            connect.execute("CREATE TABLE "+c_id+" (classid varchar(255),name varchar(255),roll varchar(255),personid varchar(255));")
            connect.commit()
            connect.close()
            os.system("python create_person_group.py "+c_id)
            return 'class added successfully'+'''<br><a href='/c_add'>Click here to go back</a>'''
    else:
        return render_template('add_class.html')
                        
                        
@app.route('/exchange',methods=['POST','GET'])
def exchange():
    if 'userid' in session:
        userid=session['userid']
    else:
        return redirect(url_for('displayNames'))
    if request.method=='POST':
        date1=currentDate
        classid=request.form['class_id']
        period=request.form['period']
        ex_sub_code=request.form['ex_sub_code']
        wb = load_workbook(filename = "F:/Autoattendance-Cognitive-master-Copy/excel/"+classid+".xlsx")
        if current_mon in wb.sheetnames:
            sheet = wb.get_sheet_by_name(current_mon)
        for i in range(1,sheet.max_column + 1):
            col = get_column_letter(i)
            if sheet['%s%s'% (col,'1')].value == date1:
                col=get_column_letter(i+int(period)-1)
                break
        sheet['%s%s' %(col,'2')].value=ex_sub_code
        wb.save(filename = "F:/Autoattendance-Cognitive-master-Copy/excel/"+classid+".xlsx")
        return redirect(url_for('view'))
    else:
        return redirect(url_for('view'))
    
@app.route("/edit",methods=["GET","POST"])
def edit():
    if 'userid' in session:
        userid=session['userid']
    else:
        return redirect(url_for('displayNames'))
    if request.method=='POST':
        roll_num=request.form['roll']
        period=request.form['period']
        class_id1=request.form['classid']
        if len(class_id1)>7:
            course_id=class_id1[:6]
            class_id=class_id1[7:]
        else:
            class_id=class_id1
        def getDateColumn():
            for i in range(1,sheet.max_column + 1):
                col = get_column_letter(i)
                if sheet['%s%s'% (col,'1')].value == currentDate:
                    return i
        wb = load_workbook(filename ="F:/Autoattendance-Cognitive-master-Copy/excel/"+class_id+".xlsx")
        current_mon=time.strftime("%B-%Y")
        if current_mon in wb.sheetnames:
            sheet = wb.get_sheet_by_name(current_mon)
        col = getDateColumn()
    
        for row in range(3,sheet.max_row + 1):
                rn = sheet['A%s'% row].value
                if rn==int(roll_num):
                    sheet['%s%s' % (get_column_letter(col+int(period)-1),str(row))]='OD'
                    wb.save(filename ="F:/Autoattendance-Cognitive-master-Copy/excel/"+class_id+".xlsx")        
        if len(class_id1)>7:
    
            wb = xlrd.open_workbook("F:/Autoattendance-Cognitive-master-Copy/excel/"+class_id+'.xlsx')
            current_mon=time.strftime("%B-%Y")
            sh = wb.sheet_by_name(current_mon)

          #  os.remove("E:/Autoattendance-Cognitive-master-Copy/csv/"+course_id+"_"+class_id+'.csv')
            your_csv_file = open("F:/Autoattendance-Cognitive-master-Copy/csv/"+course_id+"_"+class_id+'.csv', 'w', encoding='utf8')
            wr = csv.writer(your_csv_file, quoting=csv.QUOTE_ALL)

            for rownum in range(sh.nrows):
                wr.writerow(sh.row_values(rownum))

            your_csv_file.close()
            m=course_id
            def csv_to_html_table(fname,headers=None,delimiter=","):
                with open(fname) as f:
                    content = f.readlines()
                    #reading file content into list
                    rows = [x.strip() for x in content]
                    table = "<table border='1'>"
                    #creating HTML header row if header is provided 
                    if headers is not None:
                        table+= "".join(["<th>"+cell[1:-1]+"</th>" for cell in headers.split(delimiter)])
                    else:
                        table+= "".join(["<th>"+cell[1:-1]+"</th>" for cell in rows[0].split(delimiter)])
                        rows=rows[1:]
                    #Converting csv to html row by row
                    i=0
                    index=[0,1]
                    for cell in rows[1].split(delimiter):
                        if cell[1:-1]==m:
                            index.append(i)
                        i=i+1
                    rows=rows[3::2]
                    for row in rows:
                        row=row.split(delimiter)
                        table+="<tr>"
                        for i in index:
                            data=row[i]
                            table+="".join(["<td>"+data[1:-1]+"</td>"])
                        table+="</tr>" + "\n"
                    table+="</table><br>"
                    return table

            x=csv_to_html_table("F:/Autoattendance-Cognitive-master-Copy/csv/"+course_id+"_"+class_id+'.csv')
          #  os.remove("E:/Autoattendance-Cognitive-master-Copy/templates/"+course_id+"_"+class_id+'.html')
            f=open("F:/Autoattendance-Cognitive-master-Copy/templates/"+course_id+"_"+class_id+".html",'w')
            f.write(x)
            y='''<form action="{{url_for('edit')}}" method='POST'><input type="text" name="classid" value="'''+course_id+"_"+class_id+'''"readonly><br>
Date :'''+currentDate+'''<br> Roll num:<input type="text" name="roll"><br>   Period Num :<input type="text" name="period" ><br>Attendance : On Duty <br> <input type="submit" value="submit" name="submit">
</form><br>
<br>
<body>
<button onclick="myFunction()">Print this page</button>

<script>
function myFunction() {
    window.print();
}
</script>
</body>
<form action="{{url_for('logout')}}">
<input type="submit" name="Logout" value="Logout">
</form><a href="/view">Go back</a>'''
            f.write(y)
            f.close()
            return render_template(course_id+"_"+class_id+".html")
           
        else:
        
            wb = xlrd.open_workbook("F:/Autoattendance-Cognitive-master-Copy/excel/"+class_id+'.xlsx')
            current_mon=time.strftime("%B-%Y")
            sh = wb.sheet_by_name(current_mon)
           # os.remove("E:/Autoattendance-Cognitive-master-Copy/csv/"+class_id+'.csv')
            your_csv_file = open("F:/Autoattendance-Cognitive-master-Copy/csv/"+class_id+'.csv', 'w', encoding='utf8')
            wr = csv.writer(your_csv_file, quoting=csv.QUOTE_ALL)

            for rownum in range(sh.nrows):
                wr.writerow(sh.row_values(rownum))

            your_csv_file.close()

            def csv_to_html_table_main(fname,headers=None,delimiter=","):
                with open(fname) as f:
                    content = f.readlines()
                    #reading file content into list
                    rows = [x.strip() for x in content]
                    #print(rows)
                    table = "<table border='1'>"
                    #creating HTML header row if header is provided 
                    if headers is not None:
                        table+= "".join(["<th>"+cell[1:-1]+"</th>" for cell in headers.split(delimiter)])
                    else:
                        table+= "".join(["<th>"+cell[1:-1]+"</th>" for cell in rows[0].split(delimiter)])
                        rows=rows[1:]
                    #Converting csv to html row by row
                    for row in rows:
                        table+= "<tr>" + "".join(["<td>"+cell[1:-1]+"</td>" for cell in row.split(delimiter)]) + "</tr>" + "\n"
                    table+="</table><br>"
                    return table

            x=csv_to_html_table_main("F:/Autoattendance-Cognitive-master-Copy/csv/"+class_id+'.csv')
           # os.remove("E:/Autoattendance-Cognitive-master-Copy/templates/"+class_id+'.html')
            f=open("F:/Autoattendance-Cognitive-master-Copy/templates/"+class_id+".html",'w')
            f.write(x)
            y='''<form action="{{url_for('edit')}}" method='POST'><input type="text" name="classid" value="'''+class_id+'''"readonly><br>
Date :'''+currentDate+'''<br> Roll num:<input type="text" name="roll"><br>   Period Num :<input type="text" name="period" ><br>Attendance : On Duty <br> <input type="submit" value="submit" name="submit">
</form><br>
<br>
<button onclick="myFunction()">Print this page</button>
<script>
function myFunction() {
    window.print();
}
</script>
<form action="{{url_for('logout')}}">
<input type="submit" name="Logout" value="Logout">
</form><a href="/view">Go back</a>'''
            f.write(y)
            #print(x)
            f.close()
            return render_template(class_id+'.html')

    else:
        return redirect(url_for('view'))
    
@app.route("/logout")
def logout():
    session.pop('userid',None)
    return redirect(url_for('index'))

@app.before_request
def make_session_permanent():
    session.permanent = True
    app.permanent_session_lifetime = timedelta(minutes=60)
    
if __name__ == '__main__':
    app.config['TEMPLATES_AUTO_RELOAD']=True
    app.run()  #this starts the server at IP address 127.0.0.1 and port 5000. if you want to start it on a remote IP address just use app.run("IP_ADDRESS", "Port-no")
