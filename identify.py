import cognitive_face as CF
import os,pathlib
import sqlite3,sys
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import time,datetime
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

classid="IT2015"#sys.argv[1]
personGroupId=classid.lower()
#get current date
currentDate = time.strftime("%d_%m_%y")
wb = load_workbook(filename = "F:/Autoattendance-Cognitive-master-Copy/excel/"+classid+".xlsx")
d = datetime.datetime.today()
current_mon=d.strftime("%B-%Y")
sheet = wb[current_mon]

current_time=d.strftime("%H%M")
x='none'
if '0800'<current_time and current_time<'0850':
    x='one'
elif '0851'<current_time and current_time<'0950':
    x='two'
elif'1006'<current_time and current_time<'1055':
    x='three'
elif'1056'<current_time and current_time<'1145':
    x='four'
elif'1146'<current_time and current_time<'1235':
    x='five'
elif'1320'<current_time and current_time<'1400':
    x='six'
elif'1401'<current_time and current_time<'1510':
    x='seven'
#print(x)
    
def getDateColumn():
    for i in range(1,sheet.max_column + 1):
        col = get_column_letter(i)
        if sheet['%s%s'% (col,'1')].value == currentDate:
            return i

col = getDateColumn()
attend=[]			
Key = 'ec5f3015dee74f15a8c864ab1783b740'
CF.Key.set(Key)
BASE_URL = 'https://westcentralus.api.cognitive.microsoft.com/face/v1.0/'
CF.BaseUrl.set(BASE_URL)

connect = connect = sqlite3.connect("Face-DataBase")
c = connect.cursor()

attend = []	

currentDir = os.path.dirname(os.path.abspath(__file__))
directory = os.path.join(currentDir, 'Cropped_faces\\'+classid+"\\")
print(directory)
for filename in os.listdir(directory):
    if filename.endswith(".jpg"):  
        imgurl = pathlib.Path(os.path.join(directory, filename)).as_uri() 
        res = CF.face.detect(imgurl[8:])
        #print(res)
        if len(res) != 1:
            print ("No face detected.")
            continue
        faceIds = []
        for face in res:
            faceIds.append(face['faceId'])
        res = CF.face.identify(faceIds, personGroupId)
        for face  in res:
                    if not face['candidates']:
                        print ("Unknown")
                    else:
                        personId = face['candidates'][0]['personId']
                        c.execute("SELECT * FROM "+classid+" WHERE personID = ?", (personId,))
                        row = c.fetchone()
                       # if row is not None:
                        x1=row[2]
                        attend.append(x1)
                        print (row[1] + " recognized")
                        time.sleep(6)		
                        #print(attend)
for row in range(3,sheet.max_row + 1):
    rn = sheet['A%s'% row].value
    rn=(str(rn))
    #print(type(rn))
    if rn is not None:   
        for y in attend:
     #       print(type(y))
            if rn==y:
                #print(rn)
                if x=='one':
                    sheet['%s%s' % (get_column_letter(col),str(row))]='1'
                    #print('%s%s' % (get_column_letter(col),str(row)))
                elif x=='two':
                    sheet['%s%s' % (get_column_letter(col+1),str(row))]='1'
                    #print('%s%s' % (get_column_letter(col+1),str(row)))
                elif x=='three':
                    sheet['%s%s' % (get_column_letter(col+2),str(row))]='1'
                    #print('%s%s' % (get_column_letter(col+2),str(row)))
                elif x=='four':
                    sheet['%s%s' % (get_column_letter(col+3),str(row))]='1'
                    #print('%s%s' % (get_column_letter(col+3),str(row)))
                elif x=='five':
                    sheet['%s%s' % (get_column_letter(col+4),str(row))]='1'
                    #print('%s%s' % (get_column_letter(col+4),str(row)))
                elif x=='six':
                    sheet['%s%s' % (get_column_letter(col+5),str(row))]='1'
                    #print('%s%s' % (get_column_letter(col+5),str(row)))                    
                elif x=='seven':
                    sheet['%s%s' % (get_column_letter(col+6),str(row))]='1'
                    #print('%s%s' % (get_column_letter(col+6),str(row)))
                #wb.save(filename = "E:/Autoattendance-Cognitive-master-Copy/excel/"+classid+".xlsx")

for row in range(3,sheet.max_row + 1):
    rn = sheet['A%s'% row].value
    rn=(str(rn))
    #print(type(rn))
    if rn is not None:
                if x=='one' and sheet['%s%s' % (get_column_letter(col),str(row))].value is None:
                    sheet['%s%s' % (get_column_letter(col),str(row))]='AB'
                    #print('%s%s' % (get_column_letter(col),str(row)))
                elif x=='two' and sheet['%s%s' % (get_column_letter(col+1),str(row))].value is None:
                    sheet['%s%s' % (get_column_letter(col+1),str(row))]='AB'
                    #print('%s%s' % (get_column_letter(col+1),str(row)))
                elif x=='three' and sheet['%s%s' % (get_column_letter(col+2),str(row))].value is None:
                    sheet['%s%s' % (get_column_letter(col+2),str(row))]='AB'
                    #print('%s%s' % (get_column_letter(col+2),str(row)))
                elif x=='four' and sheet['%s%s' % (get_column_letter(col+3),str(row))].value is None:
                    sheet['%s%s' % (get_column_letter(col+3),str(row))]='AB'
                    #print('%s%s' % (get_column_letter(col+3),str(row)))
                elif x=='five' and sheet['%s%s' % (get_column_letter(col+4),str(row))].value is None:
                    sheet['%s%s' % (get_column_letter(col+4),str(row))]='AB'
                    #print('%s%s' % (get_column_letter(col+4),str(row)))
                elif x=='six' and sheet['%s%s' % (get_column_letter(col+5),str(row))].value is None:
                    sheet['%s%s' % (get_column_letter(col+5),str(row))]='AB'
                    #print('%s%s' % (get_column_letter(col+5),str(row)))                    
                elif x=='seven' and sheet['%s%s' % (get_column_letter(col+6),str(row))].value is None:
                    sheet['%s%s' % (get_column_letter(col+6),str(row))]='AB'
                    #print(rn)
                    #print('%s%s' % (get_column_letter(col+6),str(row)))                
wb.save(filename = "F:/Autoattendance-Cognitive-master-Copy/excel/"+classid+".xlsx")	 	
#currentDir = os.path.dirname(os.path.abspath(__file__))
#imgurl = urllib.pathname2url(os.path.join(currentDir, "1.jpg"))
#res = CF.face.detect(imgurl)
#faceIds = []
#for face in res:
 #   faceIds.append(face['faceId'])

#res = CF.face.identify(faceIds,personGroupId)
# for face in res:
#     personName = CF.person.get(personGroupId, face['candidates']['personId'])
#     print personName
#print res